import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from database.repositories import bom_repo, compatibility_repo, vehicle_repo, node_materials_repo

# Lv1 명칭 → 8대 분류 코드 자동 감지
_CATEGORY_KEYWORDS = {
    "1": ["전력추진", "주전력"],
    "2": ["연결"],
    "3": ["보조전원"],
    "4": ["운전실", "제어"],
    "5": ["제동"],
    "6": ["주행", "대차"],
    "7": ["차상신호", "신호"],
    "8": ["차체", "설비", "내외"],
}


def _detect_category(name: str) -> str:
    if not name:
        return "1"
    for code, keywords in _CATEGORY_KEYWORDS.items():
        if any(kw in name for kw in keywords):
            return code
    return "1"

CATEGORY_NAMES = {
    "1": "전력추진",
    "2": "연결",
    "3": "보조전원",
    "4": "운전실및제어",
    "5": "제동",
    "6": "주행",
    "7": "차상신호",
    "8": "차체및차내외설비",
}

BOM_COLUMNS = [
    "LEVEL", "PATH", "NODE_TYPE", "CATEGORY_CODE", "CATEGORY_NAME",
    "MATERIAL_NO", "CORP_MATERIAL_NO", "NAME", "NAME_EN", "SPECIFICATION",
    "QUANTITY", "UNIT", "MANUFACTURER", "MANUFACTURER_PN",
    "DRAWING_NO", "WEIGHT_KG", "MATERIAL", "NOTES", "INDENT_NAME",
]

COMPAT_COLUMNS = [
    "MATERIAL_NO", "VEHICLE_CODE", "VEHICLE_NAME", "COMPAT_TYPE", "NOTES",
]


def _header_style():
    return {
        "fill": PatternFill("solid", fgColor="1F4E79"),
        "font": Font(bold=True, color="FFFFFF"),
        "alignment": Alignment(horizontal="center", vertical="center"),
    }


def _category_style():
    return {
        "fill": PatternFill("solid", fgColor="D6E4F0"),
        "font": Font(bold=True),
    }


def _apply_style(cell, style: dict):
    for attr, val in style.items():
        setattr(cell, attr, val)


def export_bom(vehicle_type_id: int) -> bytes:
    nodes = bom_repo.get_tree(vehicle_type_id)
    vehicle = vehicle_repo.get_vehicle_by_id(vehicle_type_id)

    wb = Workbook()

    # ── 시트 1: BOM_DATA ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "BOM_DATA"
    ws.freeze_panes = "A2"

    header_style = _header_style()
    for col_idx, col_name in enumerate(BOM_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        _apply_style(cell, header_style)

    for row_idx, node in enumerate(nodes, 2):
        depth = node.get("depth", 0)
        category_code = node.get("category_code", "")
        is_category = node.get("node_type") == "category"

        # 새 테이블에서 공사 자재번호 목록 조회 (세미콜론 구분)
        mats = node_materials_repo.get_by_node(node["id"])
        corp_nos = "; ".join(m["corp_material_no"] for m in mats) if mats else node.get("corp_material_no", "") or ""

        values = {
            "LEVEL": depth + 1,
            "PATH": node.get("path", ""),
            "NODE_TYPE": node.get("node_type", ""),
            "CATEGORY_CODE": category_code,
            "CATEGORY_NAME": CATEGORY_NAMES.get(str(category_code), ""),
            "MATERIAL_NO": node.get("material_no", ""),
            "CORP_MATERIAL_NO": corp_nos,
            "NAME": node.get("name", ""),
            "NAME_EN": node.get("name_en", ""),
            "SPECIFICATION": node.get("specification", ""),
            "QUANTITY": node.get("quantity", 1),
            "UNIT": node.get("unit", "EA"),
            "MANUFACTURER": node.get("manufacturer", ""),
            "MANUFACTURER_PN": node.get("manufacturer_pn", ""),
            "DRAWING_NO": node.get("drawing_no", ""),
            "WEIGHT_KG": node.get("weight_kg", ""),
            "MATERIAL": node.get("material", ""),
            "NOTES": node.get("notes", ""),
            "INDENT_NAME": "  " * depth + node.get("name", ""),
        }

        for col_idx, col_name in enumerate(BOM_COLUMNS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=values[col_name])
            if is_category:
                _apply_style(cell, _category_style())

    # 열 너비 자동 조정
    col_widths = {
        "LEVEL": 6, "PATH": 12, "NODE_TYPE": 12, "CATEGORY_CODE": 6,
        "CATEGORY_NAME": 14, "MATERIAL_NO": 16, "CORP_MATERIAL_NO": 14,
        "NAME": 30, "NAME_EN": 24,
        "SPECIFICATION": 20, "QUANTITY": 8, "UNIT": 6, "MANUFACTURER": 16,
        "MANUFACTURER_PN": 16, "DRAWING_NO": 16, "WEIGHT_KG": 10,
        "MATERIAL": 12, "NOTES": 20, "INDENT_NAME": 40,
    }
    for col_idx, col_name in enumerate(BOM_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 14)

    # ── 시트 2: COMPATIBILITY ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("COMPATIBILITY")
    for col_idx, col_name in enumerate(COMPAT_COLUMNS, 1):
        cell = ws2.cell(row=1, column=col_idx, value=col_name)
        _apply_style(cell, header_style)

    # part/kit 노드의 호환성 정보 수집
    row_idx = 2
    for node in nodes:
        if node.get("node_type") not in ("part", "kit"):
            continue
        compat_list = compatibility_repo.get_by_node(node["id"])
        for c in compat_list:
            ws2.cell(row=row_idx, column=1, value=node.get("material_no", ""))
            ws2.cell(row=row_idx, column=2, value=c.get("vehicle_code", ""))
            ws2.cell(row=row_idx, column=3, value=c.get("vehicle_name", ""))
            ws2.cell(row=row_idx, column=4, value=c.get("compat_type", "compatible"))
            ws2.cell(row=row_idx, column=5, value=c.get("notes", ""))
            row_idx += 1

    for col_idx, width in enumerate([16, 14, 16, 14, 20], 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def get_template() -> bytes:
    """빈 BOM 템플릿 Excel 반환"""
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM_DATA"
    ws.freeze_panes = "A2"
    header_style = _header_style()
    for col_idx, col_name in enumerate(BOM_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        _apply_style(cell, header_style)
    # 예시 행
    examples = [
        (1, "000010", "category", "1", "전력추진", "", "전력추진", "", "", 1, "EA", "", "", "", "", "", "", "전력추진"),
        (2, "000010.000010", "assembly", "1", "전력추진", "KTX-001", "주변압기조립체", "Main Transformer Assy", "", 1, "SET", "", "", "", "", "", "", "  주변압기조립체"),
        (3, "000010.000010.000010", "part", "1", "전력추진", "KTX-001-A", "주변압기", "Main Transformer", "25kV/1MVA", 1, "EA", "현대로템", "HRT-MT-001", "", 4500.0, "", "", "    주변압기"),
    ]
    for row_idx, ex in enumerate(examples, 2):
        for col_idx, val in enumerate(ex, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    ws2 = wb.create_sheet("COMPATIBILITY")
    for col_idx, col_name in enumerate(COMPAT_COLUMNS, 1):
        cell = ws2.cell(row=1, column=col_idx, value=col_name)
        _apply_style(cell, header_style)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def import_bom(vehicle_type_id: int, file_bytes: bytes) -> dict:
    """
    Excel → DB import.
    반환: {"inserted": N, "updated": N, "errors": [...]}
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    if "BOM_DATA" not in wb.sheetnames:
        return {"inserted": 0, "updated": 0, "errors": ["BOM_DATA 시트를 찾을 수 없습니다."]}

    ws = wb["BOM_DATA"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_map = {h: i for i, h in enumerate(headers) if h}

    def get_val(row, col_name):
        idx = col_map.get(col_name)
        if idx is None:
            return None
        return row[idx].value

    inserted = 0
    updated = 0
    errors = []

    # 기존 노드를 material_no로 인덱싱 (upsert용)
    existing = {
        n["material_no"]: n
        for n in bom_repo.get_tree(vehicle_type_id)
        if n.get("material_no")
    }

    # LEVEL 스택으로 parent_id 추적
    id_stack: dict[int, int] = {}  # level → last created id

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        name = get_val(row, "NAME")
        if not name:
            continue

        level = get_val(row, "LEVEL")
        try:
            level = int(level) if level is not None else 1
        except (ValueError, TypeError):
            level = 1

        material_no = get_val(row, "MATERIAL_NO") or None
        node_type = get_val(row, "NODE_TYPE") or "assembly"
        category_code = str(get_val(row, "CATEGORY_CODE")) if get_val(row, "CATEGORY_CODE") else None

        parent_id = id_stack.get(level - 1) if level > 1 else None

        data = {
            "parent_id": parent_id,
            "vehicle_type_id": vehicle_type_id,
            "node_type": node_type,
            "category_code": category_code,
            "material_no": material_no,
            "name": str(name),
            "name_en": get_val(row, "NAME_EN"),
            "specification": get_val(row, "SPECIFICATION"),
            "unit": get_val(row, "UNIT") or "EA",
            "quantity": float(get_val(row, "QUANTITY") or 1),
            "manufacturer": get_val(row, "MANUFACTURER"),
            "manufacturer_pn": get_val(row, "MANUFACTURER_PN"),
            "drawing_no": get_val(row, "DRAWING_NO"),
            "weight_kg": float(get_val(row, "WEIGHT_KG")) if get_val(row, "WEIGHT_KG") else None,
            "material": get_val(row, "MATERIAL"),
            "notes": get_val(row, "NOTES"),
            "sort_order": None,
        }

        try:
            if material_no and material_no in existing:
                node = bom_repo.update_node(existing[material_no]["id"], data)
                updated += 1
            else:
                node = bom_repo.create_node(data)
                if material_no:
                    existing[material_no] = node
                inserted += 1
            id_stack[level] = node["id"]
        except Exception as e:
            errors.append(f"행 {row_idx}: {e}")

    return {"inserted": inserted, "updated": updated, "errors": errors}


# ──────────────────────────────────────────────────────────────────────────────
# 네이티브 포맷 (BOM 그려본 자료.xlsx 형식) import
# 컬럼 구조: L1~L10 | LEVEL | Lv1~Lv10 | 영문 명칭 | 원제작사 OEM Part NO. |
#            비고 | 분류 코드 | [타차종1] | [타차종2] ...
# ──────────────────────────────────────────────────────────────────────────────

def preview_native(file_bytes: bytes) -> dict:
    """파일 내 시트 목록과 각 시트의 예상 행 수 반환"""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    result = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 데이터행 수 추정 (빈 행 제외)
        count = sum(
            1 for row in ws.iter_rows(min_row=4, values_only=True)
            if any(v is not None for v in row)
        )
        result.append({"sheet": sheet_name, "rows": count})
    wb.close()
    return {"sheets": result}


def _parse_native_headers(ws):
    """
    3행 헤더에서 컬럼 인덱스 맵 반환.
    반환: {
        'l': [0-based idx for L1..L10],
        'lv': [0-based idx for Lv1..Lv10],
        'level': idx, 'name_en': idx, 'cpn': idx,
        'notes': idx, 'class_code': idx,
        'other_vehicles': [(col_name, idx), ...]
    }
    """
    headers = [cell.value for cell in next(ws.iter_rows(min_row=3, max_row=3))]

    def find(name):
        try:
            return headers.index(name)
        except ValueError:
            return None

    l_idx = [find(f"L{i}") for i in range(1, 11)]
    lv_idx = [find(f"Lv{i}") for i in range(1, 11)]
    level_idx = find("LEVEL")
    name_en_idx = find("영문 명칭")
    # OEM Part NO. 헤더는 줄바꿈 포함
    cpn_idx = next((i for i, h in enumerate(headers) if h and "OEM Part NO" in str(h)), None)
    notes_idx = find("비고")
    class_idx = find("분류 코드")

    # 타차종 열: L1~분류코드 이후에 오는 텍스트 열
    fixed_names = {f"L{i}" for i in range(1, 11)} | {f"Lv{i}" for i in range(1, 11)} | {
        "LEVEL", "영문 명칭", "비고", "분류 코드", None
    }
    other_vehicles = []
    for i, h in enumerate(headers):
        if h and str(h).strip() and h not in fixed_names and "OEM" not in str(h):
            # class_idx 이후 열만 타차종으로 간주
            if class_idx is not None and i > class_idx:
                other_vehicles.append((str(h).strip(), i))

    return {
        "l": l_idx, "lv": lv_idx, "level": level_idx,
        "name_en": name_en_idx, "cpn": cpn_idx,
        "notes": notes_idx, "class_code": class_idx,
        "other_vehicles": other_vehicles,
    }


def import_native_bom(
    vehicle_type_id: int,
    file_bytes: bytes,
    sheet_name: str,
    other_vehicle_map: dict[str, int],   # {"이음": vehicle_type_id, ...}
) -> dict:
    """
    네이티브 BOM 포맷 import.

    other_vehicle_map: 타차종 열 이름 → vehicle_type_id 매핑
                       (없으면 호환성 등록 생략)

    반환: {"inserted": N, "updated": N, "compat_added": N, "errors": [...]}
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    if sheet_name not in wb.sheetnames:
        return {"inserted": 0, "updated": 0, "compat_added": 0,
                "errors": [f"'{sheet_name}' 시트를 찾을 수 없습니다."]}

    ws = wb[sheet_name]
    h = _parse_native_headers(ws)

    inserted = 0
    updated = 0
    compat_added = 0
    errors = []

    # 분류코드 기준으로 기존 노드 인덱싱 (upsert용)
    existing_by_class: dict[str, dict] = {}
    for n in bom_repo.get_tree(vehicle_type_id):
        if n.get("material_no"):
            existing_by_class[n["material_no"]] = n

    # level → node_id 스택 (부모 추적)
    id_stack: dict[int, int] = {}
    # Lv1 이름 → category_code 캐시
    lv1_category: dict[str, str] = {}

    for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), 4):
        # 빈 행 스킵
        if all(v is None for v in row):
            continue

        # 레벨 읽기
        level_raw = row[h["level"]] if h["level"] is not None else None
        if level_raw is None:
            continue
        try:
            level = int(float(str(level_raw)))
        except (ValueError, TypeError):
            continue
        if level < 1 or level > 10:
            continue

        # 자재내역: Lv{level} 열
        lv_col = h["lv"][level - 1]
        name = row[lv_col] if lv_col is not None else None
        if not name:
            continue
        name = str(name).strip()

        # 영문명, 제작사 CPN, 비고, 분류코드
        name_en = str(row[h["name_en"]]).strip() if h["name_en"] is not None and row[h["name_en"]] else None
        this_cpn = str(row[h["cpn"]]).strip() if h["cpn"] is not None and row[h["cpn"]] else None
        notes = str(row[h["notes"]]).strip() if h["notes"] is not None and row[h["notes"]] else None
        class_code = str(row[h["class_code"]]).strip() if h["class_code"] is not None and row[h["class_code"]] else None

        # Lv1 기반 카테고리 코드 결정
        lv1_col = h["lv"][0]
        lv1_name = str(row[lv1_col]).strip() if lv1_col is not None and row[lv1_col] else None
        if lv1_name and lv1_name not in lv1_category:
            lv1_category[lv1_name] = _detect_category(lv1_name)
        category_code = lv1_category.get(lv1_name or "", "1")

        # 노드 타입 결정
        if level == 1:
            node_type = "category"
        else:
            node_type = "assembly"  # 나중에 자식 없으면 part로 간주해도 되나, 일단 assembly

        parent_id = id_stack.get(level - 1) if level > 1 else None

        data = {
            "parent_id": parent_id,
            "vehicle_type_id": vehicle_type_id,
            "node_type": node_type,
            "category_code": category_code,
            "material_no": class_code,   # 분류코드를 임시 식별자로 활용
            "name": name,
            "name_en": name_en,
            "specification": None,
            "unit": "EA",
            "quantity": 1,
            "manufacturer": None,
            "manufacturer_pn": this_cpn,
            "drawing_no": None,
            "weight_kg": None,
            "material": None,
            "notes": notes,
            "sort_order": None,
        }

        try:
            if class_code and class_code in existing_by_class:
                node = bom_repo.update_node(existing_by_class[class_code]["id"], data)
                updated += 1
            else:
                node = bom_repo.create_node(data)
                if class_code:
                    existing_by_class[class_code] = node
                inserted += 1

            id_stack[level] = node["id"]

            # ── 타차종 호환성 등록 ─────────────────────────────────────
            if this_cpn and other_vehicle_map:
                for col_name, col_idx in h["other_vehicles"]:
                    other_vid = other_vehicle_map.get(col_name)
                    if other_vid is None:
                        continue
                    other_cpn = str(row[col_idx]).strip() if row[col_idx] else None
                    if not other_cpn:
                        continue  # None → 호환성 미등록
                    compat_type = "compatible" if other_cpn == this_cpn else "partial"
                    compatibility_repo.upsert(node["id"], other_vid, compat_type,
                                              notes=f"CPN: {other_cpn}" if compat_type == "partial" else None)
                    compat_added += 1

        except Exception as e:
            errors.append(f"행 {row_idx}: {e}")

    return {"inserted": inserted, "updated": updated,
            "compat_added": compat_added, "errors": errors}

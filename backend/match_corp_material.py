"""
자재마스터 엑셀에서 공사 자재번호를 BOM 노드에 매핑 시도.
매칭 방법:
  1) CPN 완전일치 (manufacturer_pn ↔ 제조자 부품번호)
  2) CPN ↔ 자재번호 내역에 포함된 코드 (내역 안에 CPN/도면번호가 박혀있는 경우)
  3) 도면번호 ↔ 자재번호 내역에 포함된 코드
  4) 명칭 정규화 완전일치
"""
import sys, re, sqlite3
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

MASTER_PATH = r'C:\Users\shg98\Desktop\철도공사\1차 백데이터\자재마스터 260415.XLSX'
DB_PATH = r'C:\Dev\railway-bom\backend\data\bom.db'

def normalize(s):
    if not s: return ''
    s = str(s).strip()
    if ':' in s:
        s = s.rsplit(':', 1)[-1]
    s = re.sub(r'[\s\-_/,\.\(\)\[\]·×]', '', s)
    return s.lower()

# 내역 문자열에서 코드처럼 보이는 토큰 추출
# 영문+숫자 혼합, 5자 이상
CODE_RE = re.compile(r'[A-Za-z][A-Za-z0-9\-\.]{4,}')

def extract_codes_from_name(name: str) -> list[str]:
    if not name: return []
    # 콜론 뒤 실제 품명 부분에서 추출
    part = name.rsplit(':', 1)[-1] if ':' in name else name
    return [m.group() for m in CODE_RE.finditer(part)]

print("자재마스터 로딩 중...")
wb = openpyxl.load_workbook(MASTER_PATH, read_only=True, data_only=True)
ws = wb.active
headers = [str(c) if c else '' for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
idx = {h: i for i, h in enumerate(headers)}

masters = []
for row in ws.iter_rows(min_row=2, values_only=True):
    corp_no  = str(row[idx['자재번호']]).strip() if row[idx['자재번호']] else ''
    raw_name = str(row[idx['자재번호 내역']]).strip() if row[idx['자재번호 내역']] else ''
    cpn      = str(row[idx['제조자 부품번호']]).strip() if row[idx['제조자 부품번호']] else ''
    if not corp_no or corp_no == 'None': continue
    norm = normalize(raw_name)
    embedded_codes = extract_codes_from_name(raw_name)
    masters.append({
        'corp_no': corp_no,
        'name': raw_name,
        'cpn': cpn,
        'norm': norm,
        'codes': embedded_codes,  # 내역에 박힌 코드들
    })

print(f"자재마스터 {len(masters):,}건 로드 완료")

# 인덱스 구성
cpn_map: dict[str, list] = {}       # CPN → master rows
code_map: dict[str, list] = {}      # 내역 속 코드 → master rows
name_map: dict[str, list] = {}      # 정규화 명칭 → master rows

for m in masters:
    if m['cpn']:
        cpn_map.setdefault(m['cpn'], []).append(m)
    for code in m['codes']:
        code_map.setdefault(code, []).append(m)
    if m['norm']:
        name_map.setdefault(m['norm'], []).append(m)

print("BOM DB 로딩 중...")
con = sqlite3.connect(DB_PATH)
con.row_factory = sqlite3.Row

nodes = con.execute("""
    SELECT id, name, material_no, manufacturer_pn, drawing_no, node_type
    FROM bom_nodes
    WHERE (corp_material_no IS NULL OR corp_material_no = '')
      AND node_type != 'repair_kit'
""").fetchall()
print(f"미매핑 BOM 노드 {len(nodes):,}건\n")

results = []
matched_ids = set()

for node in nodes:
    nid   = node['id']
    name  = node['name'] or ''
    mat   = node['material_no'] or ''
    cpn   = (node['manufacturer_pn'] or '').strip()
    drw   = (node['drawing_no'] or '').strip()
    norm  = normalize(name)
    candidates = []
    method = ''

    # 1) CPN 완전일치
    if cpn and cpn in cpn_map:
        candidates = cpn_map[cpn]
        method = 'CPN일치'

    # 2) CPN이 자재번호 내역 속 코드와 일치
    elif cpn and cpn in code_map:
        candidates = code_map[cpn]
        method = 'CPN↔내역코드'

    # 3) 도면번호가 자재번호 내역 속 코드와 일치
    elif drw and drw in code_map:
        candidates = code_map[drw]
        method = '도면번호↔내역코드'

    # 4) 도면번호가 CPN으로 등록된 경우
    elif drw and drw in cpn_map:
        candidates = cpn_map[drw]
        method = '도면번호=CPN'

    # 5) 명칭 완전일치
    elif norm and norm in name_map:
        candidates = name_map[norm]
        method = '명칭일치'

    if candidates:
        matched_ids.add(nid)
        for c in candidates[:3]:  # 최대 3개
            results.append({
                'method':    method,
                'bom_code':  mat,
                'bom_name':  name,
                'bom_cpn':   cpn,
                'bom_drw':   drw,
                'corp_no':   c['corp_no'],
                'corp_name': c['name'],
                'corp_cpn':  c['cpn'],
            })

con.close()

# 방법별 집계
from collections import Counter
method_counts = Counter(r['method'] for r in results)

print(f"총 매칭 노드: {len(matched_ids)}건 / 총 결과행: {len(results)}건")
for m, c in method_counts.most_common():
    print(f"  {m}: {c}건")

OUT = r'C:\Dev\railway-bom\backend\match_corp_material_result.txt'
with open(OUT, 'w', encoding='utf-8') as f:
    f.write(f"공사 자재번호 매핑 가능 후보\n")
    f.write(f"총 매칭 노드: {len(matched_ids)}건 / 결과행: {len(results)}건\n")
    for m, c in method_counts.most_common():
        f.write(f"  {m}: {c}건\n")
    f.write("=" * 110 + "\n\n")

    for method in ['CPN일치', 'CPN↔내역코드', '도면번호↔내역코드', '도면번호=CPN', '명칭일치']:
        group = [r for r in results if r['method'] == method]
        if not group: continue
        f.write(f"\n【{method}】 {len(group)}건\n")
        f.write("-" * 110 + "\n")
        for r in group:
            f.write(f"BOM: {r['bom_code']:<30s} {r['bom_name']:<35s} CPN:{r['bom_cpn']:<20s} DRW:{r['bom_drw']}\n")
            f.write(f"  → 공사번호: {r['corp_no']:<12s} | CPN:{r['corp_cpn']:<20s} | {r['corp_name']}\n\n")

print(f"\n결과 저장 완료: {OUT}")
